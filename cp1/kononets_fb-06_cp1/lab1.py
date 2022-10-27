from math import log2
from openpyxl import Workbook

alphabet_with_gap = 'абвгдеёжзийклмнопрстуфхцчшщъыьэюя '
alphabet = 'абвгдеёжзийклмнопрстуфхцчшщъыьэюя'
wb = Workbook()


def without_punctuation(string, value):             # прибираємо усі зайві знаки
    text_punctuation = ""
    if value == 1:
        text_punctuation = alphabet_with_gap
    elif value == 0:
        text_punctuation = alphabet
    for p in string:
        if p not in text_punctuation:
            # банальна заміна символа у строчці
            string = string.replace(p, '')
    while string.count("  "):
        string = string.replace("  ", ' ')
    return string


def stripped_lines(text, value):                   # робимо єдиний текст якщо текст починається з нової строчки
    with open(text, "r", encoding="utf-8") as file:
        newline_breaks = ""
        for line in file:
            if value == 1:                         # ставимо пробіл?
                stripped_line = line.strip()
            elif value == 0:                        # не ставимо пробіл?
                stripped_line = line.strip()
            newline_breaks += stripped_line.lower()
        return newline_breaks


def pretty_text(text, value):                       # приводимо текст до потрібного нам за завданням
    if value == 1:          # ставимо пробіл?
        newline_breaks = stripped_lines(text, value)
        newline_breaks = without_punctuation(newline_breaks, value)

        return newline_breaks
    elif value == 0:        # не ставимо пробіл?
        newline_breaks = stripped_lines(text, value)
        newline_breaks = without_punctuation(newline_breaks, value)
        while newline_breaks.count("  "):
            newline_breaks = newline_breaks.replace("  ", ' ')
        return newline_breaks


def frequency_of_letters(text, my_alphabet):    # рахуємо частоту зустрічаємості літер
    letters_count = {}
    for letter in my_alphabet:
        letters_count[letter] = 0               # для кожного символу(ключа) з алфавіту ставимо значення 0
    for letter in text:
        letters_count[letter] += 1              # скільки разів зустрічається стільки раз і додаємо

    letters_frequency = {}
    for letter in my_alphabet:
        letters_frequency[letter] = round((letters_count[letter])/len(text), 9)  # округлюю до 9 знаків для точності

    return letters_frequency


def bigram_frequency(text, my_alphabet, boolean_value):     # частота зустрічаємості біграм
    bigram_count = {}
    bigram_frequency_is = {}
    for letter_1 in my_alphabet:
        for letter_2 in my_alphabet:
            if letter_1+letter_2 != "  ":
                dict_key = letter_1 + letter_2
                bigram_count[dict_key] = 0

    if boolean_value == 1:          # H1
        i = 0
        while i < len(text) - 1:            # довжина тексту - 1 , для H1
            key = text[i] + text[i+1]       # скріплюю літери біграм
            bigram_count[key] = bigram_count[key] + 1   # підрахунок зустрічаємості кожної біграми
            i = i + 1

        for key in bigram_count.keys():
            bigram_frequency_is[key] = round(bigram_count[key]/(len(text)-1), 9)  # округлюю до 9 знаків для точності

    else:   # H2
        if len(text) % 2 == 1:   # якщо довжина тексту парна, то разглядаємо весь текст
            k = 1
        else:                    # у іншому випадку одиноку літеру в кінці тексту не розглядвємо
            k = 0
        i = 0
        while i < len(text) - k:            # усе інше аналогічно до H1
            key = text[i] + text[i+1]
            bigram_count[key] = bigram_count[key] + 1
            i = i + 2                       # окрім цієї частини, бо крок 2

        for key in bigram_count.keys():
            bigram_frequency_is[key] = round(bigram_count[key]/(len(text)//2), 9)
    return bigram_frequency_is


def entropy(dictionary, n):                 # знаходимо ентропію за формулою з методички(де n - це n-грамма)
    entropy_list = []
    for key in dictionary.keys():
        if dictionary[key] != 0:
            entropy_list.append(float(dictionary[key]) * log2(dictionary[key])/n)
    result = - sum(entropy_list)
    return result


def redundancy_of_language(found_entropy, my_alphabet):      # знаючи ентропію, обчислюємо надлишковість
    return 1 - (found_entropy/log2(len(my_alphabet)))        # формула з методички


def make_xl_file(dictionary, str_sheet_name, str_file_path, sheet_number):
    # 'D:\\Python\\PycharmProjects\\kononets_fb-06_cp1\\kononets_fb-06_cp1\\cp_lab1.xlsx'
    ws1 = wb.create_sheet(str_sheet_name, sheet_number)
    keys = list(dictionary.keys())
    values = list(dictionary.values())
    for i in range(1, len(keys)+1):
        ws1["A" + str(i)] = keys[i-1]
    for i in range(1, len(keys)+1):
        ws1["B" + str(i)] = values[i-1]
    wb.save(str_file_path)


def add_notes_to_xl(str_note, value, column_row_for_note, column_row_for_val, str_sheet_name, str_file_path):
    ws = wb[str_sheet_name]
    ws[column_row_for_note] = str_note
    ws[column_row_for_val] = value
    wb.save(str_file_path)


# Розкоментувати виводи нижче, якщо потрібно (уся інформація і так буде у табличці cp_lab1.xlsx)
# print("---------------- Починаємо дослід для алфавіту з пробілом ----------------")
txt_file_path = "D:\\Python\\PycharmProjects\\crypto-22-23\\cp1\\kononets_fb-06_cp1\\some_text.txt"
my_text = pretty_text(txt_file_path, True)
file_path = "D:\\Python\\PycharmProjects\\crypto-22-23\\cp1\\kononets_fb-06_cp1\\cp_lab1.xlsx"
experiment_1 = frequency_of_letters(my_text, alphabet_with_gap)
# print(experiment_1)
make_xl_file(experiment_1, "experiment_1", file_path, 0)
en_1 = entropy(experiment_1, 1)
# print(en_1)
add_notes_to_xl("Ентропія:", en_1, "D2", "E2", "experiment_1", file_path)
rud_1 = redundancy_of_language(en_1, alphabet_with_gap)
# print(rud_1)
add_notes_to_xl("Надлишковість:", rud_1, "D3", "E3", "experiment_1", file_path)

experiment_H1_1 = bigram_frequency(my_text, alphabet_with_gap, True)
# print(experiment_H1_1)
make_xl_file(experiment_H1_1, "experiment_H1_1", file_path, 1)
en_H1_1 = entropy(experiment_H1_1, 2)
# print(en_H1_1)
add_notes_to_xl("Ентропія:", en_H1_1, "D2", "E2", "experiment_H1_1", file_path)
rud_H1_1 = redundancy_of_language(en_H1_1, alphabet_with_gap)
# print(rud_H1_1)
add_notes_to_xl("Надлишковість:", rud_H1_1, "D3", "E3", "experiment_H1_1", file_path)

experiment_H2_1 = bigram_frequency(my_text, alphabet_with_gap, False)
# print(experiment_H2_1)
make_xl_file(experiment_H2_1, "experiment_H2_1", file_path, 2)
en_H2_1 = entropy(experiment_H2_1, 2)
# print(en_H2_1)
add_notes_to_xl("Ентропія:", en_H2_1, "D2", "E2", "experiment_H2_1", file_path)
rud_H2_1 = redundancy_of_language(en_H2_1, alphabet_with_gap)
# print(rud_H2_1)
add_notes_to_xl("Надлишковість:", rud_H2_1, "D3", "E3", "experiment_H2_1", file_path)

# print("---------------- Починаємо дослід для алфавіту без пробіла ----------------")
my_text_2 = pretty_text(txt_file_path, False)
experiment_2 = frequency_of_letters(my_text_2, alphabet)
# print(experiment_2)
make_xl_file(experiment_2, "experiment_2", file_path, 3)
en_2 = entropy(experiment_2, 1)
# print(en_2)
add_notes_to_xl("Ентропія:", en_2, "D2", "E2", "experiment_2", file_path)
rud_2 = redundancy_of_language(en_2, alphabet)
# print(rud_2)
add_notes_to_xl("Надлишковість:", rud_2, "D3", "E3", "experiment_2", file_path)

experiment_H1_2 = bigram_frequency(my_text_2, alphabet, True)
# print(experiment_H1_2)
make_xl_file(experiment_H1_2, "experiment_H1_2", file_path, 4)
en_H1_2 = entropy(experiment_H1_2, 2)
# print(en_H1_2)
add_notes_to_xl("Ентропія:", en_H1_2, "D2", "E2", "experiment_H1_2", file_path)
rud_H1_2 = redundancy_of_language(en_H1_2, alphabet)
# print(rud_H1_2)
add_notes_to_xl("Надлишковість:", rud_H1_2, "D3", "E3", "experiment_H1_2", file_path)

experiment_H2_2 = bigram_frequency(my_text_2, alphabet, False)
# print(experiment_H2_2)
# print(len(experiment_H2_2))
make_xl_file(experiment_H2_2, "experiment_H2_2", file_path, 5)
en_H2_2 = entropy(experiment_H2_2, 2)
# print(en_H2_2)
add_notes_to_xl("Ентропія:", en_H2_2, "D2", "E2", "experiment_H2_2", file_path)
rud_H2_2 = redundancy_of_language(en_H2_2, alphabet)
# print(rud_H2_2)
add_notes_to_xl("Надлишковість:", rud_H2_2, "D3", "E3", "experiment_H2_2", file_path)
